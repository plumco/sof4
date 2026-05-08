python

import streamlit as st
import pandas as pd
import io
import datetime
import base64 as _b64
import tempfile as _tmp
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ─── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Huliot SO Form",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  .main-header {
    background: linear-gradient(90deg, #1a3a6b 0%, #2563b0 100%);
    padding: 18px 24px; border-radius: 10px; margin-bottom: 18px;
  }
  .main-header h1 { color: white; margin: 0; font-size: 1.6rem; }
  .main-header p  { color: #cce0ff; margin: 0; font-size: 0.88rem; }
  .section-title {
    color: #1a3a6b; font-weight: 700; font-size: 0.95rem;
    border-bottom: 2px solid #2563b0; padding-bottom: 6px; margin-bottom: 12px;
  }
  .metric-box { background:#1a3a6b;color:white;border-radius:8px;padding:14px 18px;text-align:center; }
  .metric-box .val { font-size:1.6rem;font-weight:700; }
  .metric-box .lbl { font-size:0.78rem;opacity:0.8; }
  .stButton > button { background:#2563b0;color:white;border-radius:6px;border:none;padding:8px 20px; }
  .stButton > button:hover { background:#1a3a6b; }
  div[data-testid="stVerticalBlock"] > div:empty { display:none!important; }
  .stElementContainer:empty { display:none!important; }
</style>
""", unsafe_allow_html=True)

# ─── Embedded Excel ─────────────────────────────────────────────────────────────
_EXCEL_B64 = None   # Filled below at runtime from file

@st.cache_resource
def _get_excel_path():
    # Try local file first (works locally and on Streamlit Cloud if in repo)
    import os
    local = os.path.join(os.path.dirname(__file__), "SO_FORMAT_PROJECT_-_APRIL_2026.xlsx")
    if os.path.exists(local):
        return local
    # Fallback: decode from embedded base64
    if _EXCEL_B64:
        t = _tmp.NamedTemporaryFile(delete=False, suffix=".xlsx")
        t.write(_b64.b64decode(_EXCEL_B64))
        t.flush(); t.close()
        return t.name
    raise FileNotFoundError("Excel master file not found. Add SO_FORMAT_PROJECT_-_APRIL_2026.xlsx to repo root.")

EXCEL_FILE = _get_excel_path()

# ─── Product Line → Sub Group mapping (from actual Product Master sheet) ────────
# Sub groups found in data: HT PRO, US SINGLE STACK, US DOUBLE STACK,
# PERT-AL-PERT, KLIMAPRESS, ROMAFASER, ROMAKLIMA, RED FIRE PRESS, PRESS INOX, etc.
PRODUCT_LINES = {
    "HT Pro":          ["HT PRO", "HTPRO", "HT-PRO"],
    "Ultra Silent":    ["US S/S", "US D/S", "US SIN", "US DOU", "ULTRA SILENT", "ULTRA-SILENT"],
    "PERT-AL-PERT":    ["PERT", "KLIMAPRESS", "MLCP"],
    "PPR / Heliroma":  ["ROMAFASER", "ROMAKLIMA", "RED FIRE", "PRESS INOX", "PPR"],
    "Accessories":     ["CLAMP", "TOOL", "LUBRIC", "GASKET", "GRATING", "CHANNEL",
                        "DWC", "CALIBR", "SHOWER"],
}

# Fitting type → keywords in description/category
FITTING_TYPES = {
    "All":         None,
    "Pipe":        ["PIPE"],
    "Bend":        ["BEND", "ELBOW"],
    "Branch":      ["TEE", " Y ", "BRANCH", "DOUBLE Y", "CORNER"],
    "Trap":        ["TRAP", "SMART LOCK", "NHANI"],
    "Coupler":     ["COUPLER", "REPAIR", "JOINT", "LOCK SEAL", "END LOCK", "WC CONN", "WC CON"],
    "Reducer":     ["REDUCER", "ECCENTRIC", "CONCENTRIC"],
    "Inspection":  ["CLEANING", "ACCESS", "RISER", "BOSS", "VENT", "HAFF", "INSP"],
    "Clamp":       ["CLAMP"],
    "Accessory":   ["GASKET", "RING", "FLANGE", "CAP", "GRATING", "LUBRIC", "TOOL",
                    "CALIBR", "DWC", "CHANNEL"],
}

# DN sets per product line
DN_MAP = {
    "HT Pro":        [40, 50, 75, 110, 125, 160, 200],
    "Ultra Silent":  [50, 75, 110, 125, 160, 200],
    "PERT-AL-PERT":  [16, 20, 25, 32, 40, 50, 63],
    "PPR / Heliroma":[20, 25, 32, 40, 50, 63, 75, 90, 110],
    "Accessories":   [],
    "All":           [40, 50, 75, 110, 125, 160, 200],
}

# ─── Helper filters using str.contains (Arrow-safe) ────────────────────────────
def _kw_mask(series, kws):
    pat = "|".join(re.escape(k) for k in kws)
    return series.astype(str).str.contains(pat, case=False, na=False, regex=True)

def filter_by_line(df, line_key):
    if line_key == "All":
        return df
    kws = PRODUCT_LINES[line_key]
    return df[_kw_mask(df["sub_group"], kws) | _kw_mask(df["category"], kws)]

def filter_by_type(df, type_key):
    if type_key == "All" or FITTING_TYPES[type_key] is None:
        return df
    kws = FITTING_TYPES[type_key]
    return df[_kw_mask(df["category"], kws) | _kw_mask(df["description"], kws)]

# ─── Data Loaders ───────────────────────────────────────────────────────────────
@st.cache_data
def load_product_master():
    df = pd.read_excel(EXCEL_FILE, sheet_name="Product Master", header=0)
    df.columns = ["item_code","description","dn","moq","list_price",
                  "category","sub_group","hsn_code","gst_rate","cat2"]
    df = df.dropna(subset=["item_code"])
    return df

@st.cache_data
def load_validations():
    df = pd.read_excel(EXCEL_FILE, sheet_name="Data Validations", header=0)
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    return df

def _icol(df, name):
    df = df.copy()
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    return [str(v) for v in df[name].dropna().tolist()] if name in df.columns else []

products_df  = load_product_master()
val_df       = load_validations()

sales_persons    = _icol(val_df, "Sales Person")
payment_terms    = _icol(val_df, "Payment Terms")
project_types    = _icol(val_df, "Project Type")
customer_types   = _icol(val_df, "Customer Type")
freight_terms    = _icol(val_df, "Freight Terms")
order_validities = _icol(val_df, "Order Validity")
_cd_df = val_df.copy()
_cd_df.columns = [c.strip() if isinstance(c, str) else c for c in _cd_df.columns]
cd_options = sorted([round(v, 2) for v in _cd_df["CD %"].dropna().tolist()])

# ─── Session State ──────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "line_items": [],
        "so_no": "", "po_no": "",
        "po_date": datetime.date.today(),
        "order_date": datetime.date.today(),
        "sales_person": sales_persons[2] if len(sales_persons) > 2 else "",
        "business_type": "Project",
        "business_segment": project_types[1] if len(project_types) > 1 else "",
        "customer_po_no": "", "bitrix_id": "",
        "customer_type": "Builder",
        "rera_no": "", "order_validity": "7 Days", "other_remark": "",
        "cash_discount": 0.03, "freight": "FOR", "payment_term": "Advance",
        "bill_name": "", "bill_address": "", "bill_city": "", "bill_pin": "", "bill_gstin": "",
        "ship_name": "", "ship_address": "", "ship_city": "", "ship_pin": "", "ship_gstin": "",
        "consultant_name": "", "consultant_contact": "",
        "developer_name": "", "site_name": "", "other_remark2": "",
        "contact_person": "", "contact_mobile": "",
        # catalog filter state
        "cat_dn": "ALL", "cat_type": "All",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>📋 Huliot Pipes & Fittings — Sales Order Form</h1>
  <p>Format No: HPF-01 &nbsp;|&nbsp; Rev. 01 &nbsp;|&nbsp; Issue Date: 10.04.2026</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["📝 Order Details", "🛍️ Add Products", "📥 Summary & Download"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — ORDER DETAILS
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="section-title">📌 Order Identity</div>', unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4)
    with c1: st.session_state.so_no       = st.text_input("Sales Order No", value=st.session_state.so_no)
    with c2: st.session_state.po_no       = st.text_input("PO No", value=st.session_state.po_no)
    with c3: st.session_state.po_date     = st.date_input("PO Date", value=st.session_state.po_date)
    with c4: st.session_state.order_date  = st.date_input("Order Date", value=st.session_state.order_date)
    st.markdown("---")

    st.markdown('<div class="section-title">👤 Sales & Project Info</div>', unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    with c1:
        idx = sales_persons.index(st.session_state.sales_person) if st.session_state.sales_person in sales_persons else 0
        st.session_state.sales_person      = st.selectbox("Sales Person", sales_persons, index=idx)
    with c2:
        bt_opts = ["Project","Retail"]
        bt_idx  = bt_opts.index(st.session_state.business_type) if st.session_state.business_type in bt_opts else 0
        st.session_state.business_type     = st.selectbox("Business Type", bt_opts, index=bt_idx)
    with c3:
        seg_idx = project_types.index(st.session_state.business_segment) if st.session_state.business_segment in project_types else 0
        st.session_state.business_segment  = st.selectbox("Business Segment", project_types, index=seg_idx)

    c1,c2,c3,c4 = st.columns(4)
    with c1: st.session_state.rera_no        = st.text_input("RERA Project No *", value=st.session_state.rera_no)
    with c2:
        ct_idx = customer_types.index(st.session_state.customer_type) if st.session_state.customer_type in customer_types else 0
        st.session_state.customer_type     = st.selectbox("Customer Type", customer_types, index=ct_idx)
    with c3: st.session_state.customer_po_no = st.text_input("Customer PO No", value=st.session_state.customer_po_no)
    with c4: st.session_state.bitrix_id      = st.text_input("Bitrix Deal ID", value=st.session_state.bitrix_id)

    c1,c2,c3 = st.columns(3)
    with c1: st.session_state.contact_person = st.text_input("Contact Person", value=st.session_state.contact_person)
    with c2: st.session_state.contact_mobile = st.text_input("Contact Mobile", value=st.session_state.contact_mobile)
    with c3:
        ov_idx = order_validities.index(st.session_state.order_validity) if st.session_state.order_validity in order_validities else 0
        st.session_state.order_validity    = st.selectbox("Order Validity", order_validities, index=ov_idx)
    st.markdown("---")

    st.markdown('<div class="section-title">💰 Commercial Terms</div>', unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    with c1:
        cd_idx = cd_options.index(st.session_state.cash_discount) if st.session_state.cash_discount in cd_options else 0
        st.session_state.cash_discount = st.selectbox("Cash Discount %", cd_options, index=cd_idx,
                                                       format_func=lambda x: f"{x*100:.0f}%")
    with c2:
        ft_idx = freight_terms.index(st.session_state.freight) if st.session_state.freight in freight_terms else 0
        st.session_state.freight       = st.selectbox("Freight / Delivery", freight_terms, index=ft_idx)
    with c3:
        pt_idx = payment_terms.index(st.session_state.payment_term) if st.session_state.payment_term in payment_terms else 0
        st.session_state.payment_term  = st.selectbox("Payment Term", payment_terms, index=pt_idx)
    st.markdown("---")

    st.markdown('<div class="section-title">🏢 Billing & Shipping</div>', unsafe_allow_html=True)
    col_bill, col_ship = st.columns(2)
    with col_bill:
        st.markdown("**Receiver (Bill To)**")
        st.session_state.bill_name    = st.text_input("Name",    value=st.session_state.bill_name,    key="bn")
        st.session_state.bill_address = st.text_area("Address", value=st.session_state.bill_address, key="ba", height=70)
        c1,c2,c3 = st.columns(3)
        with c1: st.session_state.bill_city  = st.text_input("City",     value=st.session_state.bill_city,  key="bc")
        with c2: st.session_state.bill_pin   = st.text_input("Pin Code", value=st.session_state.bill_pin,   key="bp")
        with c3: st.session_state.bill_gstin = st.text_input("GSTIN",    value=st.session_state.bill_gstin, key="bg")
    with col_ship:
        st.markdown("**Consignee (Ship To)**")
        st.session_state.ship_name    = st.text_input("Name",    value=st.session_state.ship_name,    key="sn")
        st.session_state.ship_address = st.text_area("Address", value=st.session_state.ship_address, key="sa", height=70)
        c1,c2,c3 = st.columns(3)
        with c1: st.session_state.ship_city  = st.text_input("City",     value=st.session_state.ship_city,  key="sc")
        with c2: st.session_state.ship_pin   = st.text_input("Pin Code", value=st.session_state.ship_pin,   key="sp")
        with c3: st.session_state.ship_gstin = st.text_input("GSTIN",    value=st.session_state.ship_gstin, key="sg")
    st.markdown("---")

    st.markdown('<div class="section-title">📎 Additional Info</div>', unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        st.session_state.consultant_name    = st.text_input("Consultant / Architect Firm", value=st.session_state.consultant_name)
        st.session_state.developer_name     = st.text_input("Developer / Bungalow Owner",  value=st.session_state.developer_name)
    with c2:
        st.session_state.consultant_contact = st.text_input("Consultant Contact No.", value=st.session_state.consultant_contact)
        st.session_state.site_name          = st.text_input("Site / Project Name",    value=st.session_state.site_name)
    st.session_state.other_remark = st.text_area("Other Remark", value=st.session_state.other_remark, height=60)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ADD PRODUCTS
# ══════════════════════════════════════════════════════════════════════════════
with tab2:

    def add_item(row, disc_val):
        """Add or update a line item."""
        cd        = st.session_state.cash_discount
        net_disc  = round(1 - (1 - disc_val) * (1 - cd), 4)
        net_price = round(float(row["list_price"]) * (1 - net_disc), 2)
        existing  = [i for i, it in enumerate(st.session_state.line_items)
                     if it["item_code"] == row["item_code"]]
        if existing:
            st.session_state.line_items[existing[0]]["qty"] += int(row["moq"])
            st.toast(f"Qty updated: {row['item_code']}", icon="✅")
        else:
            st.session_state.line_items.append({
                "item_code":   row["item_code"],
                "description": row["description"],
                "dn":          row["dn"],
                "category":    row["category"],
                "sub_group":   row["sub_group"],
                "moq":         int(row["moq"]),
                "qty":         int(row["moq"]),
                "hsn_code":    row["hsn_code"],
                "list_price":  float(row["list_price"]),
                "disc_pct":    disc_val,
                "cash_disc":   cd,
                "net_disc":    net_disc,
                "net_price":   net_price,
            })
            st.toast(f"Added: {row['item_code']}", icon="✅")
        st.rerun()

    def render_product_grid(view_df, tab_key):
        """Render product rows with Add buttons."""
        if view_df.empty:
            st.info("No items match the selected filters.")
            return
        st.caption(f"Showing {min(len(view_df), 80)} of {len(view_df)} items")
        hdr = st.columns([3, 1.8, 0.7, 0.7, 1.0, 0.75, 1.0])
        for h, t in zip(hdr, ["Description","Sub Group","DN","MOQ","List ₹","Disc%",""]):
            h.markdown(f"<small><b>{t}</b></small>", unsafe_allow_html=True)
        for _, row in view_df.head(80).iterrows():
            cols = st.columns([3, 1.8, 0.7, 0.7, 1.0, 0.75, 1.0])
            cols[0].markdown(
                f"<small><b style='color:#1a3a6b'>{row['item_code']}</b>"
                f"<br>{str(row['description'])[:65]}</small>",
                unsafe_allow_html=True)
            cols[1].markdown(f"<small>{str(row['sub_group'])[:22]}</small>", unsafe_allow_html=True)
            cols[2].write(str(row["dn"]))
            cols[3].write(str(int(row["moq"])))
            cols[4].write(f"₹{float(row['list_price']):,.0f}")
            disc_key = f"disc_{row['item_code']}"
            if disc_key not in st.session_state:
                st.session_state[disc_key] = 0.52
            disc_val = cols[5].number_input(
                "", min_value=0.0, max_value=0.99, step=0.01,
                value=st.session_state[disc_key],
                key=f"d_{tab_key}_{row['item_code']}",
                label_visibility="collapsed", format="%.2f")
            st.session_state[disc_key] = disc_val
            if cols[6].button("➕ Add", key=f"a_{tab_key}_{row['item_code']}"):
                add_item(row, disc_val)

    # ── Product Line Tabs (from actual sub_group analysis) ──────────────────
    line_options = ["All"] + list(PRODUCT_LINES.keys())
    line_tabs    = st.tabs(line_options)

    for li, (line_tab, line_key) in enumerate(zip(line_tabs, line_options)):
        with line_tab:
            # Base data filtered by product line
            line_df = filter_by_line(products_df, line_key) if line_key != "All" else products_df

            # ── DN Size Buttons ───────────────────────────────────────────────
            dn_list = DN_MAP.get(line_key, [])

            # Dynamically get actual DNs present in this line's data
            actual_dns = sorted(
                [int(x) for x in line_df["dn"].dropna().unique()
                 if str(x).isdigit() or (isinstance(x, float) and not str(x).endswith('.0') == False)],
                key=lambda x: x
            )
            # Use intersection of DN_MAP + actual data (show only what exists)
            if dn_list:
                show_dns = [d for d in dn_list if any(
                    abs(float(x) - d) < 0.5
                    for x in line_df["dn"].dropna()
                    if str(x).replace('.','',1).isdigit()
                )]
            else:
                show_dns = sorted(set(
                    int(float(x)) for x in line_df["dn"].dropna()
                    if str(x).replace('.','',1).isdigit()
                ))

            dn_state_key = f"cat_dn_{li}"
            if dn_state_key not in st.session_state:
                st.session_state[dn_state_key] = "ALL"

            if show_dns:
                st.markdown(
                    "<p style='font-size:0.78rem;font-weight:700;color:#555;"
                    "text-transform:uppercase;letter-spacing:1px;margin:0 0 6px'>Select Size (DN)</p>",
                    unsafe_allow_html=True)
                btn_cols = st.columns(len(show_dns) + 1)
                all_cnt = len(line_df)
                if btn_cols[0].button(
                    f"ALL  {all_cnt}",
                    key=f"dn_ALL_{li}",
                    type="primary" if st.session_state[dn_state_key] == "ALL" else "secondary"
                ):
                    st.session_state[dn_state_key] = "ALL"; st.rerun()
                for ci, dn in enumerate(show_dns):
                    cnt = len(line_df[
                        line_df["dn"].apply(
                            lambda x: str(x).replace('.0','') == str(dn)
                        )
                    ])
                    if btn_cols[ci+1].button(
                        f"DN{dn}  {cnt}",
                        key=f"dn_{dn}_{li}",
                        type="primary" if st.session_state[dn_state_key] == str(dn) else "secondary"
                    ):
                        st.session_state[dn_state_key] = str(dn); st.rerun()
                st.markdown("")

            # ── Fitting Type Pills ────────────────────────────────────────────
            type_state_key = f"cat_type_{li}"
            if type_state_key not in st.session_state:
                st.session_state[type_state_key] = "All"

            type_icons = {
                "All":"🔵","Pipe":"□","Bend":"↩","Branch":"⌥","Trap":"⊔",
                "Coupler":"○","Reducer":"◁","Inspection":"⊙","Clamp":"∩","Accessory":"⚙"
            }
            pill_cols = st.columns(len(FITTING_TYPES))
            for ti, tkey in enumerate(FITTING_TYPES.keys()):
                ico  = type_icons.get(tkey, "•")
                active = st.session_state[type_state_key] == tkey
                if pill_cols[ti].button(
                    f"{ico} {tkey}",
                    key=f"ft_{tkey}_{li}",
                    type="primary" if active else "secondary"
                ):
                    st.session_state[type_state_key] = tkey; st.rerun()

            # ── Search bar ────────────────────────────────────────────────────
            search_key = f"cat_search_{li}"
            if search_key not in st.session_state:
                st.session_state[search_key] = ""
            search_q = st.text_input(
                "search", value=st.session_state[search_key],
                key=f"cs_{li}", label_visibility="collapsed",
                placeholder="🔍  Search item code, description, size...")
            st.session_state[search_key] = search_q
            st.markdown("")

            # ── Apply all filters ─────────────────────────────────────────────
            view_df = line_df.copy()

            # DN filter
            sel_dn = st.session_state[dn_state_key]
            if sel_dn != "ALL" and show_dns:
                view_df = view_df[
                    view_df["dn"].apply(lambda x: str(x).replace('.0','') == sel_dn)
                ]

            # Type filter
            view_df = filter_by_type(view_df, st.session_state[type_state_key])

            # Text search
            if search_q:
                sq = search_q.lower()
                view_df = view_df[
                    view_df["item_code"].astype(str).str.lower().str.contains(sq, na=False) |
                    view_df["description"].astype(str).str.lower().str.contains(sq, na=False) |
                    view_df["dn"].astype(str).str.lower().str.contains(sq, na=False)
                ]

            render_product_grid(view_df, f"{li}")

    st.markdown("---")

    # ── Bulk Paste ────────────────────────────────────────────────────────────
    with st.expander("📋 Bulk Add — Paste from Excel", expanded=False):
        st.caption("Fill Item Code + Qty below → click Import")
        _blank = [{"Item Code": "", "Qty": 1} for _ in range(10)]
        if "paste_grid" not in st.session_state:
            st.session_state.paste_grid = _blank.copy()
        edited = st.data_editor(
            st.session_state.paste_grid, num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Item Code": st.column_config.TextColumn("Item Code", width="medium"),
                "Qty":       st.column_config.NumberColumn("Qty", min_value=1, step=1, width="small"),
            },
            key="paste_editor", height=280)
        st.session_state.paste_grid = edited
        col_imp, col_clr = st.columns([1, 4])
        if col_imp.button("⬆️ Import", type="primary"):
            added = 0
            for row_d in edited:
                code = str(row_d.get("Item Code","")).strip()
                qty  = int(row_d.get("Qty", 1) or 1)
                if not code: continue
                match = products_df[products_df["item_code"].astype(str).str.strip() == code]
                if match.empty: continue
                row = match.iloc[0]
                disc_key = f"disc_{row['item_code']}"
                disc_val = st.session_state.get(disc_key, 0.52)
                cd        = st.session_state.cash_discount
                net_disc  = round(1 - (1 - disc_val) * (1 - cd), 4)
                net_price = round(float(row["list_price"]) * (1 - net_disc), 2)
                existing  = [i for i, it in enumerate(st.session_state.line_items)
                             if it["item_code"] == row["item_code"]]
                if existing:
                    st.session_state.line_items[existing[0]]["qty"] = qty
                else:
                    st.session_state.line_items.append({
                        "item_code": row["item_code"], "description": row["description"],
                        "dn": row["dn"], "category": row["category"], "sub_group": row["sub_group"],
                        "moq": int(row["moq"]), "qty": qty, "hsn_code": row["hsn_code"],
                        "list_price": float(row["list_price"]), "disc_pct": disc_val,
                        "cash_disc": cd, "net_disc": net_disc, "net_price": net_price,
                    })
                added += 1
            st.toast(f"Imported {added} items", icon="✅")
            st.session_state.paste_grid = _blank.copy()
            st.rerun()
        if col_clr.button("🗑 Clear Grid"):
            st.session_state.paste_grid = _blank.copy(); st.rerun()

    st.markdown("---")

    # ── Line Items Table ──────────────────────────────────────────────────────
    st.markdown('<div class="section-title">🛒 Line Items</div>', unsafe_allow_html=True)
    if not st.session_state.line_items:
        st.info("No products added yet.")
    else:
        h = st.columns([0.4, 2.8, 0.8, 1.2, 0.8, 1.0, 1.0, 1.2, 0.5])
        for col, lbl in zip(h, ["#","Description","DN","Category","MOQ","Qty","List ₹","Net ₹",""]):
            col.markdown(f"**{lbl}**")
        to_remove = []
        for idx, item in enumerate(st.session_state.line_items):
            cols = st.columns([0.4, 2.8, 0.8, 1.2, 0.8, 1.0, 1.0, 1.2, 0.5])
            cols[0].write(str(idx+1))
            cols[1].markdown(f"<small><b>{item['item_code']}</b><br>{item['description'][:50]}</small>",
                             unsafe_allow_html=True)
            cols[2].write(str(item["dn"]))
            cols[3].markdown(f"<small>{item['category'][:20]}</small>", unsafe_allow_html=True)
            cols[4].write(str(item["moq"]))
            new_qty = cols[5].number_input("q", min_value=0, step=1, value=item["qty"],
                                            key=f"qty_{idx}", label_visibility="collapsed")
            if new_qty != item["qty"]:
                st.session_state.line_items[idx]["qty"] = new_qty
                cd = st.session_state.cash_discount
                nd = round(1 - (1 - item["disc_pct"]) * (1 - cd), 4)
                st.session_state.line_items[idx]["net_price"] = round(item["list_price"] * (1 - nd), 2)
            cols[6].write(f"₹{item['list_price']:,.0f}")
            cols[7].write(f"₹{item['net_price']:,.2f}")
            if cols[8].button("🗑", key=f"rem_{idx}"): to_remove.append(idx)
        for i in reversed(to_remove):
            st.session_state.line_items.pop(i)
        if to_remove: st.rerun()

        total_basic = sum(i["qty"] * i["net_price"] for i in st.session_state.line_items)
        total_qty   = sum(i["qty"] for i in st.session_state.line_items)
        st.markdown("---")
        m1,m2,m3 = st.columns(3)
        m1.markdown(f'<div class="metric-box"><div class="val">{len(st.session_state.line_items)}</div><div class="lbl">Products</div></div>', unsafe_allow_html=True)
        m2.markdown(f'<div class="metric-box"><div class="val">{total_qty}</div><div class="lbl">Total Qty</div></div>', unsafe_allow_html=True)
        m3.markdown(f'<div class="metric-box"><div class="val">₹{total_basic:,.0f}</div><div class="lbl">Order Value</div></div>', unsafe_allow_html=True)
        if st.button("🗑️ Clear All"):
            st.session_state.line_items = []; st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — SUMMARY & DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    s = st.session_state
    total_basic = sum(i["qty"] * i["net_price"] for i in s.line_items)

    errors = []
    if not s.so_no:        errors.append("Sales Order No is blank")
    if not s.sales_person: errors.append("Sales Person not selected")
    if not s.rera_no:      errors.append("RERA Project No is mandatory")
    if not s.bill_name:    errors.append("Bill-to Name is blank")
    if not s.line_items:   errors.append("No line items added")

    for e in errors:
        st.warning(f"⚠️ {e}")
    if not errors:
        st.success("✅ All required fields filled. Ready to download.")
    st.markdown("---")

    st.markdown('<div class="section-title">📋 Order Summary</div>', unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        st.markdown(f"""
| Field | Value |
|---|---|
| **SO No** | {s.so_no or '—'} |
| **Date** | {s.order_date} |
| **Sales Person** | {s.sales_person} |
| **Business Segment** | {s.business_segment} |
| **RERA No** | {s.rera_no or '—'} |
""")
    with c2:
        st.markdown(f"""
| Field | Value |
|---|---|
| **Cash Discount** | {s.cash_discount*100:.0f}% |
| **Payment Term** | {s.payment_term} |
| **Freight** | {s.freight} |
| **Bill To** | {s.bill_name or '—'} |
| **Order Value** | ₹{total_basic:,.2f} |
""")

    if s.line_items:
        rows = [{"#": i+1, "Item Code": it["item_code"], "Description": it["description"],
                 "DN": it["dn"], "Qty": it["qty"],
                 "List ₹": f"₹{it['list_price']:,.0f}",
                 "Disc%": f"{it['disc_pct']*100:.1f}%",
                 "Net ₹": f"₹{it['net_price']:,.2f}",
                 "Basic ₹": f"₹{it['qty']*it['net_price']:,.2f}"}
                for i, it in enumerate(s.line_items)]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown('<div class="section-title">📥 Download</div>', unsafe_allow_html=True)

    def build_excel():
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Sales Order"]

        _merge_map = {}
        for rng in ws.merged_cells.ranges:
            tl = ws.cell(rng.min_row, rng.min_col)
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    _merge_map[(row, col)] = tl

        def w(addr, val):
            c = ws[addr]
            target = _merge_map.get((c.row, c.column), c)
            if not isinstance(target, MergedCell):
                target.value = val

        # Header inputs (never touch formula cells)
        w("C6", s.business_type);     w("F6", s.sales_person)
        w("J6", s.po_no);             w("N6", s.po_date.strftime("%d.%m.%Y"))
        w("C7", s.business_segment);  w("F7", s.rera_no)
        w("J7", s.order_validity)
        w("C8", s.customer_po_no);    w("J8", s.so_no)
        w("N8", s.order_date.strftime("%d.%m.%Y"))
        w("C9", s.bitrix_id);         w("F9", s.customer_type)
        w("J9", s.contact_person);    w("N9", s.contact_mobile)
        # Bill To
        w("B11", s.bill_name);    w("I11", s.ship_name)
        w("B12", s.bill_address); w("I12", s.ship_address)
        w("B14", s.bill_city);    w("I14", s.ship_city)
        w("B15", s.bill_pin);     w("I15", s.ship_pin)
        w("B16", s.bill_gstin);   w("I16", s.ship_gstin)
        # Commercial — C17 is formula =IF(C19="Advance ",3%,0%) → only write C19
        w("C18", s.freight);      w("C19", s.payment_term)
        w("D20", s.consultant_name);  w("L20", s.consultant_contact)
        w("D21", s.developer_name);   w("L21", s.site_name)
        w("C22", s.other_remark)

        # Line items: write ONLY input cols B (item code), G (qty), J (disc%)
        # All other cols are VLOOKUP/formula — leave untouched
        for r in range(25, 100):
            for col_l in ("B", "G", "J"):
                c = ws[f"{col_l}{r}"]
                if not isinstance(c, MergedCell):
                    c.value = None

        for i, item in enumerate(s.line_items):
            r = 25 + i
            if r > 99: break
            w(f"B{r}", item["item_code"])
            w(f"G{r}", item["qty"])
            w(f"J{r}", round(item["disc_pct"], 4))

        # Dashboard sheet
        try:
            wd = wb["Dash Board"]
            for addr, val in [("B5", s.so_no), ("B6", s.sales_person),
                               ("B7", s.bill_name), ("B8", s.bill_city),
                               ("B9", s.bill_gstin), ("B10", s.bitrix_id),
                               ("B14", s.freight), ("B15", s.payment_term),
                               ("B16", s.rera_no)]:
                c = wd[addr]
                if not isinstance(c, MergedCell): c.value = val
        except Exception: pass

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return buf

    col_dl, col_info = st.columns([2, 3])
    with col_dl:
        if st.button("🔄 Generate Excel File", type="primary", disabled=bool(errors)):
            with st.spinner("Building Excel..."):
                try:
                    excel_bytes = build_excel()
                    fname = f"SO_{s.so_no or 'DRAFT'}_{s.order_date.strftime('%d%m%Y')}.xlsx"
                    st.download_button(
                        label="📥 Download Excel (.xlsx)",
                        data=excel_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as ex:
                    st.error(f"Error: {ex}")
    with col_info:
        st.info("💡 Downloaded file uses original HPF-01 template with formulas intact. "
                "Open in Excel → Ctrl+Alt+F9 to recalculate.")
